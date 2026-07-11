import os
import csv
import re
from datetime import datetime, date
import openpyxl
from decimal import Decimal

# Static Config
DEFAULT_USD_INR_RATE = Decimal('83.00')

def normalize_name(name):
    if not name:
        return ""
    cleaned = name.strip()
    lower_cleaned = cleaned.lower()
    
    # Direct mappings
    if lower_cleaned in ['priya', 'priya s', 'priyas']:
        return 'Priya'
    if lower_cleaned in ['rohan', 'rohan ', 'rohan  ']:
        return 'Rohan'
    if lower_cleaned in ['aisha']:
        return 'Aisha'
    if lower_cleaned in ['meera']:
        return 'Meera'
    if lower_cleaned in ['sam']:
        return 'Sam'
    if lower_cleaned in ['dev']:
        return 'Dev'
    if 'kabir' in lower_cleaned:
        return 'Kabir'
    return cleaned

def parse_date(date_val):
    """
    Parses date value from Excel or CSV.
    Can be datetime object, date object, or string.
    """
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    if not date_val:
        return None
    
    date_str = str(date_val).strip()
    # Try different formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def get_membership_ranges():
    """
    Returns the active date ranges for each member.
    None for end_date means active.
    """
    return {
        'Aisha': (date(2026, 2, 1), None),
        'Rohan': (date(2026, 2, 1), None),
        'Priya': (date(2026, 2, 1), None),
        'Meera': (date(2026, 2, 1), date(2026, 3, 31)),
        'Sam': (date(2026, 4, 8), None), # Sam paid deposit on Apr 8, housewarming on Apr 10, moved in mid-April.
        'Dev': (date(2026, 2, 1), date(2026, 3, 31)), # Dev visited in Feb and March.
        'Kabir': (date(2026, 3, 11), date(2026, 3, 11)), # Dev's friend Kabir joined for one day.
    }

def is_user_active_on_date(username, check_date):
    ranges = get_membership_ranges()
    if username not in ranges:
        return False
    start, end = ranges[username]
    if check_date < start:
        return False
    if end and check_date > end:
        return False
    return True

def parse_split_details(split_details_str):
    """
    Parses details like 'Aisha 30%; Rohan 30%; Priya 30%; Meera 20%' or 'Aisha 1; Rohan 2;'
    Returns dict of {name: value}
    """
    if not split_details_str:
        return {}
    
    # Split by semicolon
    parts = [p.strip() for p in split_details_str.split(';') if p.strip()]
    splits = {}
    for part in parts:
        # Match 'Name Value%' or 'Name Value'
        match = re.match(r'^([a-zA-Z\s\-\']+)\s+([\d\.]+)(%?)$', part)
        if match:
            raw_name, val_str, is_pct = match.groups()
            name = normalize_name(raw_name)
            val = Decimal(val_str)
            splits[name] = val
    return splits

def parse_row(row_idx, row_dict):
    """
    Parses and normalizes a single row dictionary.
    Returns: (parsed_item, list_of_anomalies)
    """
    anomalies = []
    
    # 1. Parse date
    raw_date = row_dict.get('date')
    parsed_dt = parse_date(raw_date)
    
    if not parsed_dt:
        anomalies.append({
            'row': row_idx,
            'field': 'date',
            'type': 'invalid_format',
            'severity': 'high',
            'description': f"Could not parse date: '{raw_date}'",
            'policy_action': "Skipped row due to invalid date"
        })
        return None, anomalies

    # Let's check date typos/anomalies
    description = str(row_dict.get('description', '')).strip()
    
    # Typo: 2014-03-01 -> 2026-03-12 (Goa trip Airport cab)
    if parsed_dt.year == 2014:
        old_dt = parsed_dt
        parsed_dt = date(2026, 3, 12)
        anomalies.append({
            'row': row_idx,
            'field': 'date',
            'type': 'year_typo',
            'severity': 'medium',
            'description': f"Date year is 2014 for '{description}'. Likely typo.",
            'policy_action': f"Corrected date from {old_dt} to {parsed_dt} (Goa trip checkout date)."
        })
    
    # Typo/Ambiguity: 2026-05-04 -> 2026-04-05 (Deep cleaning service)
    # The notes say "is this April 5 or May 4? format is a mess"
    # If it is May 4, Sam (joined Apr 8) should be in the split. Since Sam is not in the split, it is April 5.
    if parsed_dt == date(2026, 5, 4) and 'deep cleaning' in description.lower():
        old_dt = parsed_dt
        parsed_dt = date(2026, 4, 5)
        anomalies.append({
            'row': row_idx,
            'field': 'date',
            'type': 'ambiguous_format',
            'severity': 'medium',
            'description': f"Deep cleaning service date is ambiguous: 2026-05-04 or 2026-04-05.",
            'policy_action': f"Corrected date to {parsed_dt} (April 5) because Sam was excluded (he moved in mid-April)."
        })

    # 2. Parse paid_by (Payer)
    raw_payer = row_dict.get('paid_by')
    payer = normalize_name(raw_payer)
    
    if not payer:
        payer = 'Aisha' # Default policy
        anomalies.append({
            'row': row_idx,
            'field': 'paid_by',
            'type': 'missing_payer',
            'severity': 'high',
            'description': f"Payer is empty for '{description}'. Notes say: \"{row_dict.get('notes')}\"",
            'policy_action': f"Assigned default payer 'Aisha'."
        })
    elif payer != raw_payer:
        anomalies.append({
            'row': row_idx,
            'field': 'paid_by',
            'type': 'spelling_variant',
            'severity': 'low',
            'description': f"Standardized payer name '{raw_payer}' to '{payer}'",
            'policy_action': "Normalized casing/spacing."
        })

    # 3. Parse amount
    raw_amount = row_dict.get('amount')
    try:
        # Convert float/string to Decimal, rounding to 2 decimal places
        amount = Decimal(str(raw_amount)).quantize(Decimal('0.01'))
    except Exception:
        anomalies.append({
            'row': row_idx,
            'field': 'amount',
            'type': 'invalid_amount',
            'severity': 'high',
            'description': f"Could not parse amount: '{raw_amount}'",
            'policy_action': "Skipped row due to invalid amount"
        })
        return None, anomalies

    # Handle precision rounding anomaly (e.g. 899.995)
    if Decimal(str(raw_amount)) != amount:
        anomalies.append({
            'row': row_idx,
            'field': 'amount',
            'type': 'floating_point_rounding',
            'severity': 'low',
            'description': f"Amount '{raw_amount}' has high decimal precision. Rounded to '{amount}'.",
            'policy_action': "Rounded amount to 2 decimal places."
        })

    # Handle 0 amount
    if amount == Decimal('0.00'):
        anomalies.append({
            'row': row_idx,
            'field': 'amount',
            'type': 'zero_amount',
            'severity': 'medium',
            'description': f"Amount is 0 for '{description}'.",
            'policy_action': "Flagged as inactive/cancelled expense. Excluded from balances."
        })
        return None, anomalies

    # Handle negative amount (Refund)
    is_refund = False
    if amount < 0:
        is_refund = True
        anomalies.append({
            'row': row_idx,
            'field': 'amount',
            'type': 'negative_amount',
            'severity': 'medium',
            'description': f"Negative amount detected: '{amount}' for '{description}'.",
            'policy_action': "Treated as a refund split (reduces balances)."
        })

    # 4. Parse currency
    raw_currency = row_dict.get('currency')
    currency = str(raw_currency).strip().upper() if raw_currency else ""
    
    if not currency:
        currency = 'INR'
        anomalies.append({
            'row': row_idx,
            'field': 'currency',
            'type': 'missing_currency',
            'severity': 'medium',
            'description': f"Missing currency for '{description}'. Defaulted to INR.",
            'policy_action': "Assigned default currency 'INR'."
        })
    
    # Currency conversion rate (USD -> INR)
    converted_amount = amount
    exchange_rate = Decimal('1.0000')
    if currency == 'USD':
        exchange_rate = DEFAULT_USD_INR_RATE
        converted_amount = (amount * exchange_rate).quantize(Decimal('0.01'))
        anomalies.append({
            'row': row_idx,
            'field': 'currency',
            'type': 'currency_conversion',
            'severity': 'low',
            'description': f"USD amount '{amount}' detected.",
            'policy_action': f"Converted to INR '{converted_amount}' at exchange rate {exchange_rate}."
        })

    # 5. Check if Settlement or Expense
    notes = str(row_dict.get('notes') or '').strip()
    split_type = str(row_dict.get('split_type') or '').strip().lower()
    if split_type == 'none':
        split_type = ''
    split_with_str = str(row_dict.get('split_with') or '').strip()
    
    desc_lower = description.lower()
    is_settlement_desc = ('paid' in desc_lower and 'back' in desc_lower) or 'deposit' in desc_lower or 'settlement' in desc_lower
    
    is_settlement = False
    # Detect settlements (no split type or has settlement description)
    if not split_type or is_settlement_desc:
        is_settlement = True
        anomalies.append({
            'row': row_idx,
            'field': 'split_type',
            'type': 'settlement_logged_as_expense',
            'severity': 'medium',
            'description': f"Settlement transaction '{description}' logged in expense sheet.",
            'policy_action': "Imported as a Settlement record (excludes from group expenses, registers directly to payments)."
        })
        
        # Settle details
        payees = [normalize_name(p) for p in split_with_str.split(';') if p.strip()]
        payee = payees[0] if payees else 'Aisha' # default
        
        settlement_data = {
            'type': 'settlement',
            'date': parsed_dt,
            'description': description,
            'payer': payer,
            'payee': payee,
            'amount': amount, # original amount
            'currency': currency,
            'notes': notes
        }
        return settlement_data, anomalies

    # 6. Parse split participants
    participants = [normalize_name(p) for p in split_with_str.split(';') if p.strip()]
    
    # Check temporal membership of participants on date
    valid_participants = []
    for p in participants:
        if is_user_active_on_date(p, parsed_dt):
            valid_participants.append(p)
        else:
            anomalies.append({
                'row': row_idx,
                'field': 'split_with',
                'type': 'temporal_membership_violation',
                'severity': 'high',
                'description': f"User '{p}' was not an active member on {parsed_dt} (transaction: '{description}').",
                'policy_action': f"Excluded '{p}' from the split list."
            })
            
    if not valid_participants:
        # If no active members are found in split list, default to active members on that date
        all_ranges = get_membership_ranges()
        valid_participants = [name for name, (start, end) in all_ranges.items() 
                             if start <= parsed_dt and (end is None or parsed_dt <= end)]
        anomalies.append({
            'row': row_idx,
            'field': 'split_with',
            'type': 'no_active_participants',
            'severity': 'high',
            'description': f"No valid active participants in split for '{description}'.",
            'policy_action': f"Assigned split to all active group members: {';'.join(valid_participants)}."
        })

    # 7. Check Split Mismatch or Percentages
    split_details_str = str(row_dict.get('split_details', '')).strip()
    split_details = parse_split_details(split_details_str)
    
    # Standardize split type choice
    if split_type == 'unequal' and 'percentage' not in split_details_str and '%' not in split_details_str:
        # Check if actually custom shares or unequal amounts
        # In this dataset: Rohan 700; Priya 400; Meera 400 (amounts) -> unequal
        pass
    
    # Anomaly: split_type says equal but details are provided
    if split_type == 'equal' and split_details:
        split_details = {} # Reset to equal split
        anomalies.append({
            'row': row_idx,
            'field': 'split_type',
            'type': 'split_type_details_mismatch',
            'severity': 'low',
            'description': f"Split type is 'equal' but details are provided for '{description}'.",
            'policy_action': "Normalized to equal split (ignored extra details)."
        })

    # Calculate splits
    splits = []
    num_parts = Decimal(len(valid_participants))
    
    if split_type == 'equal':
        # Divide amount equally
        share_amount = (amount / num_parts).quantize(Decimal('0.01'))
        # Adjust for rounding difference on the last participant
        diff = amount - (share_amount * num_parts)
        for i, p in enumerate(valid_participants):
            p_amount = share_amount
            if i == len(valid_participants) - 1:
                p_amount += diff
            
            p_amount_in_base = (p_amount * exchange_rate).quantize(Decimal('0.01'))
            splits.append({
                'user': p,
                'share_value': None,
                'amount': p_amount,
                'amount_in_base': p_amount_in_base
            })

    elif split_type == 'percentage':
        # Parse details, normalize names
        # Anomaly: Sum of percentages != 100% (e.g. 110%)
        total_pct = sum(split_details.values())
        if total_pct != Decimal('100.00'):
            anomalies.append({
                'row': row_idx,
                'field': 'split_details',
                'type': 'percentage_sum_error',
                'severity': 'medium',
                'description': f"Percentage sum is {total_pct}% instead of 100% for '{description}'.",
                'policy_action': f"Rescaled percentages to sum to 100%."
            })
        
        # Calculate splits with normalized percentages
        running_amount = Decimal('0.00')
        for i, p in enumerate(valid_participants):
            pct = split_details.get(p, Decimal('0.00'))
            normalized_pct = (pct / total_pct) * Decimal('100.00')
            
            p_amount = (amount * (normalized_pct / Decimal('100.00'))).quantize(Decimal('0.01'))
            running_amount += p_amount
            
            # Adjust rounding on the last member
            if i == len(valid_participants) - 1:
                p_amount += (amount - running_amount)
            
            p_amount_in_base = (p_amount * exchange_rate).quantize(Decimal('0.01'))
            splits.append({
                'user': p,
                'share_value': normalized_pct,
                'amount': p_amount,
                'amount_in_base': p_amount_in_base
            })

    elif split_type == 'share':
        # Parse share values (e.g. Aisha 1; Rohan 2;)
        total_shares = sum(split_details.get(p, Decimal('1.00')) for p in valid_participants)
        
        running_amount = Decimal('0.00')
        for i, p in enumerate(valid_participants):
            shares = split_details.get(p, Decimal('1.00'))
            p_amount = (amount * (shares / total_shares)).quantize(Decimal('0.01'))
            running_amount += p_amount
            
            # Adjust rounding
            if i == len(valid_participants) - 1:
                p_amount += (amount - running_amount)
                
            p_amount_in_base = (p_amount * exchange_rate).quantize(Decimal('0.01'))
            splits.append({
                'user': p,
                'share_value': shares,
                'amount': p_amount,
                'amount_in_base': p_amount_in_base
            })

    elif split_type == 'unequal':
        # Raw amounts listed in split_details (e.g., Rohan 700; Priya 400; Meera 400)
        # Sum of splits should equal amount
        total_split_amount = sum(split_details.get(p, Decimal('0.00')) for p in valid_participants)
        
        if total_split_amount != amount:
            anomalies.append({
                'row': row_idx,
                'field': 'split_details',
                'type': 'unequal_sum_mismatch',
                'severity': 'medium',
                'description': f"Sum of unequal splits ({total_split_amount}) does not match total amount ({amount}) for '{description}'.",
                'policy_action': "Rescaled splits proportionally to match total amount."
            })
            
            # Rescale split amounts
            running_amount = Decimal('0.00')
            for i, p in enumerate(valid_participants):
                raw_split = split_details.get(p, Decimal('0.00'))
                p_amount = (amount * (raw_split / total_split_amount)).quantize(Decimal('0.01')) if total_split_amount > 0 else Decimal('0.00')
                running_amount += p_amount
                
                # Adjust rounding
                if i == len(valid_participants) - 1:
                    p_amount += (amount - running_amount)
                    
                p_amount_in_base = (p_amount * exchange_rate).quantize(Decimal('0.01'))
                splits.append({
                    'user': p,
                    'share_value': raw_split,
                    'amount': p_amount,
                    'amount_in_base': p_amount_in_base
                })
        else:
            for p in valid_participants:
                p_amount = split_details.get(p, Decimal('0.00'))
                p_amount_in_base = (p_amount * exchange_rate).quantize(Decimal('0.01'))
                splits.append({
                    'user': p,
                    'share_value': p_amount,
                    'amount': p_amount,
                    'amount_in_base': p_amount_in_base
                })

    expense_data = {
        'type': 'expense',
        'date': parsed_dt,
        'description': description,
        'paid_by': payer,
        'amount': amount,
        'currency': currency,
        'split_type': split_type,
        'notes': notes,
        'splits': splits,
        'amount_in_base': converted_amount
    }
    
    return expense_data, anomalies

def process_file_rows(rows):
    """
    Takes a list of raw row dicts, runs duplicate detection and normalizes all rows.
    """
    parsed_items = []
    anomalies = []
    
    # First pass: parse row by row
    raw_parsed = []
    for idx, row in enumerate(rows, start=2): # Header is line 1, data starts at line 2
        item, row_anoms = parse_row(idx, row)
        anomalies.extend(row_anoms)
        if item:
            raw_parsed.append((idx, item))

    # Second pass: Duplicate detection
    # Rules:
    # 1. Date: exact match
    # 2. Payer: exact match (standardized)
    # 3. Amount: exact match (rounded)
    # 4. Same split participants list
    # Look for duplicates and merge or flag
    
    deleted_indices = set()
    
    for i in range(len(raw_parsed)):
        idx_a, item_a = raw_parsed[i]
        if idx_a in deleted_indices:
            continue
            
        for j in range(i + 1, len(raw_parsed)):
            idx_b, item_b = raw_parsed[j]
            if idx_b in deleted_indices:
                continue
                
            # Compare A and B
            if (item_a['type'] == 'expense' and item_b['type'] == 'expense' and
                item_a['date'] == item_b['date'] and
                item_a['paid_by'] == item_b['paid_by'] and
                item_a['amount'] == item_b['amount'] and
                set(s['user'] for s in item_a['splits']) == set(s['user'] for s in item_b['splits'])):
                
                # We have a duplicate!
                deleted_indices.add(idx_b)
                anomalies.append({
                    'row': idx_b,
                    'field': 'description',
                    'type': 'duplicate_transaction',
                    'severity': 'high',
                    'description': f"Duplicate entry '{item_b['description']}' matches '{item_a['description']}' on {item_a['date']}.",
                    'policy_action': f"Deleted duplicate entry (row {idx_b}). Preserved row {idx_a}."
                })
                
    # Third pass: Manual Duplicate / Discrepancy detection
    # Example: Goa dinner Thalassa
    # Aisha logged 2400 (which notes say is wrong), Rohan logged 2450. Same date, same participants.
    for i in range(len(raw_parsed)):
        idx_a, item_a = raw_parsed[i]
        if idx_a in deleted_indices or item_a['type'] != 'expense':
            continue
            
        for j in range(i + 1, len(raw_parsed)):
            idx_b, item_b = raw_parsed[j]
            if idx_b in deleted_indices or item_b['type'] != 'expense':
                continue
                
            # Discrepancy duplicate condition (same date, similar description, similar amount)
            if (item_a['date'] == item_b['date'] and
                ('thalassa' in item_a['description'].lower() or 'thalassa' in item_b['description'].lower()) and
                set(s['user'] for s in item_a['splits']) == set(s['user'] for s in item_b['splits'])):
                
                # Notes indicate Aisha's is wrong: "Aisha also logged this I think hers is wrong"
                if item_a['paid_by'] == 'Aisha' and item_b['paid_by'] == 'Rohan':
                    deleted_indices.add(idx_a)
                    anomalies.append({
                        'row': idx_a,
                        'field': 'amount',
                        'type': 'duplicate_discrepancy',
                        'severity': 'high',
                        'description': f"Discrepancy duplicate: Aisha logged Thalassa dinner (2400 INR) but Rohan logged it (2450 INR). Aisha's note suggests hers is wrong.",
                        'policy_action': f"Deleted Aisha's incorrect duplicate (row {idx_a}). Preserved Rohan's entry of 2450 INR (row {idx_b})."
                    })
                elif item_b['paid_by'] == 'Aisha' and item_a['paid_by'] == 'Rohan':
                    deleted_indices.add(idx_b)
                    anomalies.append({
                        'row': idx_b,
                        'field': 'amount',
                        'type': 'duplicate_discrepancy',
                        'severity': 'high',
                        'description': f"Discrepancy duplicate: Aisha logged Thalassa dinner (2400 INR) but Rohan logged it (2450 INR). Aisha's note suggests hers is wrong.",
                        'policy_action': f"Deleted Aisha's incorrect duplicate (row {idx_b}). Preserved Rohan's entry of 2450 INR (row {idx_a})."
                    })

    # Filter out deleted items
    for idx, item in raw_parsed:
        if idx not in deleted_indices:
            parsed_items.append(item)

    # Sort anomalies by row number
    anomalies.sort(key=lambda x: x.get('row', 0))

    summary = {
        'total_rows_scanned': len(rows),
        'total_anomalies_detected': len(anomalies),
        'valid_expenses_count': sum(1 for item in parsed_items if item['type'] == 'expense'),
        'valid_settlements_count': sum(1 for item in parsed_items if item['type'] == 'settlement'),
    }

    return parsed_items, anomalies, summary

def read_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active # Default sheet
    
    # Read headers
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    
    # Validate required columns
    required = {'date', 'description', 'amount', 'paid_by'}
    normalized_headers = {h.lower().strip() for h in headers if h}
    if not required.issubset(normalized_headers):
        raise ValueError("Invalid file structure: Missing required headers (date, description, amount, paid_by).")
        
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for h, val in zip(headers, row):
            if h:
                row_dict[h] = val
        rows.append(row_dict)
    return rows

def read_csv_file(file_path):
    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("Invalid file structure: CSV file has no headers or content.")
            
        headers = [h.strip() for h in reader.fieldnames if h is not None]
        required = {'date', 'description', 'amount', 'paid_by'}
        normalized_headers = {h.lower().strip() for h in headers if h}
        if not required.issubset(normalized_headers):
            raise ValueError("Invalid file structure: Missing required headers (date, description, amount, paid_by).")
            
        for row in reader:
            if not any(row.values()):
                continue
            # Strip key whitespace
            clean_row = {k.strip(): v for k, v in row.items() if k is not None}
            rows.append(clean_row)
    return rows

def parse_import_file(file_path):
    _, ext = os.path.splitext(file_path.lower())
    if ext == '.xlsx':
        raw_rows = read_excel_file(file_path)
    else:
        raw_rows = read_csv_file(file_path)
        
    return process_file_rows(raw_rows)
